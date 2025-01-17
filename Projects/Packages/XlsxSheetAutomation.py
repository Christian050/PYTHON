import openpyxl
from openpyxl.chart import BarChart, Reference


class Code:
    def __init__(self, filename, sheet):
        # Sheet in Uppercase format along with number eg.'Sheet1'.
        self.filename = openpyxl.load_workbook(filename)
        self.sheet = self.filename[sheet]

    def ProcessWorkbook(self, calcRow):
        # CalcRow for cell needing calculation.
        self.calcRow = int(calcRow)

        for row in range(2, self.sheet.max_row + 1):
            self.cell = self.sheet.cell(row, self.calcRow)
            # Modify Here.
            self.cell.value = self.cell.value * 0.9
            self.productCell = self.sheet.cell(row, self.calcRow + 1)
            self.productCell.value = self.cell.value

    def DrawChart(self, minCol, maxCol):
        self.maxRow = self.sheet.max_row
        # Where the chart starts.
        self.minCol = int(minCol)
        # Where the chart ends  (duplicate as minCol if necessary).
        self.maxCol = int(maxCol)

        self.values = Reference(self.sheet,
                                min_row=2,
                                max_row=self.maxRow,
                                min_col=self.minCol,
                                max_col=self.maxCol)

        self.chart = BarChart()
        self.chart.add_data(self.values)

        self.sheet.add_chart(self.chart, 'E2')

        # Overwrites the excel file unless otherwise specified.
        self.filename.save(self.filename.filename)


# Example usage:
code = Code('example.xlsx', 'Sheet1')
code.ProcessWorkbook(3)
code.DrawChart(1, 3)
